# Automated Email Account Creation Script
# Author: Abdulrahim Yousuf
# Contact: me@obaydev1.com
# GitHub: https://github.com/z14i/automated-email-creation
#
# This Python script automates the process of creating email accounts on a cPanel interface using Selenium WebDriver.
# For usage instructions, please refer to the README.md file in the project repository.
#
import time
import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# Replace with the path to your ChromeDriver executable
webdriver_path = './chromedriver'

# Set Chrome options
chrome_options = Options()
chrome_options.add_argument('--no-sandbox')  # This is optional, depending on your setup
chrome_options.add_argument('--disable-dev-shm-usage')  # This is optional, depending on your setup

# Initialize Chrome WebDriver with options
driver = webdriver.Chrome(service=Service(executable_path=webdriver_path), options=chrome_options)

# Open the Excel file and read names
excel_file = 'names.xlsx'  # Replace with your Excel file path
df = pd.read_excel(excel_file)

# Open the URL
url = 'https://[yourdomain]:2083/cpsess2144229727/frontend/jupiter/email_accounts/index.html#/list'
driver.get(url)

# Find the input fields and enter credentials
input_field_user = driver.find_element(By.ID, 'user')
input_field_user.send_keys('omanadventures')

input_field_pass = driver.find_element(By.ID, 'pass')
input_field_pass.send_keys('7D5KZ@{af_&D')

# Click the login button
login_button = driver.find_element(By.ID, 'login_submit')
login_button.click()

# Wait for the "Create Email Account" button to be present and click it
try:
    create_button = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.ID, 'btnCreateEmailAccount'))
    )
    create_button.click()
except Exception as e:
    print(f"Error: {e}")
    driver.quit()
    exit()

# Create or open the workbook to save generated emails and passwords
output_excel_file = 'generated_emails.xlsx'
try:
    workbook = load_workbook(output_excel_file)
    sheet = workbook.active
except FileNotFoundError:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Email", "Password"])

# Iterate over names from the Excel file
for index, row in df.iterrows():
    # Wait for the "Create Email Account" button to be clickable and click it
    create_button = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.ID, 'btnCreateEmailAccount'))
    )
    create_button.click()

    # Wait for the input field to be present and interactable
    username_field = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.ID, 'txtUserName'))
    )
    WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.ID, 'txtUserName'))
    )

    # Format name: replace spaces with dots
    formatted_name = row['Name'].replace(' ', '.')

    # Scroll to the username field and input formatted name
    driver.execute_script("arguments[0].scrollIntoView(true);", username_field)
    username_field.clear()  # Clear existing text (if any)
    username_field.send_keys(formatted_name)

    # Wait for the "Generate Password" button to be clickable and click it
    generate_password_button = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.ID, 'txtEmailPassword_btnGenerate'))
    )
    generate_password_button.click()

    # Optionally, wait for a short period to ensure password is generated
    time.sleep(2)  # Adjust as needed based on your page load speed or process time

    # Retrieve the generated password from the password field
    password_field = driver.find_element(By.ID, 'txtEmailPassword')
    generated_password = password_field.get_attribute('value')

    # Click the "Create" button to submit the form
    submit_button = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.ID, 'btnCreateEmailAccount'))
    )
    submit_button.click()
    
    # Wait for 10 seconds
    time.sleep(10)

    # Save the email and generated password to the workbook
    email = f"{formatted_name}@omanadventures.com"
    sheet.append([email, generated_password])

# Save the workbook
workbook.save(output_excel_file)

# Keep the browser open for approximately 1 hour
time.sleep(3600)  # 3600 seconds = 1 hour

# Close the browser
driver.quit()
